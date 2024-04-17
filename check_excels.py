import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import matplotlib.pyplot as plt
import numpy as np

# Define the path to the original Excel file
directory_path = '/ExcelFiles'
output_directory = os.path.join(directory_path, 'duplicates')

# Function to interpolate colors from red to yellow (existing function)
def interpolate_colors_red_to_yellow(num_colors):
    colors = []
    for i in range(num_colors):
        green = int(255 * (i / (num_colors ))**0.7)  # Linear interpolation for the green channel
        colors.append(f"FF{green:02X}00")  # Format as HEX, keeping red at 255 and blue at 0
    return colors

# New function to interpolate colors from green to blue
def interpolate_colors_green_to_blue(num_colors):
    colors = []
    for i in range(num_colors):
        blue = int(255 * (i / (num_colors ))**0.7)  # Linear interpolation for the blue channel
        colors.append(f"00FF{blue:02X}")  # Format as HEX, keeping green at 255 and red at 0
    return colors

def extract_first_five_decimals(value):
    try:
        parts = str(value).split('.')
        if len(parts) > 1 and (len(parts[1]) > 5):
            significant_decimals = parts[1].lstrip('0')[:5]
            return int(significant_decimals) if significant_decimals else None
    except Exception as e:
        return None

def extract_digits(value, digit_count=6):
    try:
        parts = str(value).split('.')
        if len(parts) > 1 and (len(parts[1]) > 3):
            if parts[0] == '0':  # If zero before the decimal
                decimal_part = parts[1].lstrip('0')
            else:
                decimal_part = parts[0][-2:] + parts[1]  # Two from before the decimal and start of after
            return decimal_part[:digit_count] if len(decimal_part) >= digit_count else decimal_part
    except Exception as e:
        return None
# Function to extract the first significant digit
def extract_first_significant_digit(value):
    try:
        number_str = ''.join(filter(str.isdigit, str(value)))
        # Find the index of the first non-zero digit
        for digit in number_str:
            if digit != '0':
                return int(digit)

    except (ValueError, TypeError):
        return None

def plot_benford_law(actual_counts, total_data_points):
    digits = list(range(1, 10))
    benford_percentages = [30.1, 17.6, 12.5, 9.7, 7.9, 6.7, 5.8, 5.1, 4.6]
    expected_counts = [p * total_data_points / 100 for p in benford_percentages]

    plt.figure(figsize=(10, 5))
    plt.bar(digits, actual_counts, alpha=0.7, label='Actual Data', color='b')
    plt.plot(digits, expected_counts, 'r--', label='Expected Benford Distribution', linewidth=2)
    plt.xlabel('Digits')
    plt.ylabel('Frequency')
    plt.title('Benford\'s Law Analysis')
    plt.legend()
    plt.grid(True)
    plt.show()



def process_file(file_path, output_dir):
    file_extension = file_path.split('.')[-1]
    workbook_modified = False
    decimal_duplication_found = False
        # Set up a new workbook for each file
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    if file_extension in ['xlsx', 'xls']:
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        read_func = lambda sheet_name: pd.read_excel(xls, sheet_name=sheet_name)
    elif file_extension == 'csv':
        sheet_names = [None]  # CSV files don't have multiple sheets, but we use a list to keep the structure
        read_func = lambda _: pd.read_csv(file_path)
    elif file_extension == 'txt':
        sheet_names = [None]  # CSV files don't have multiple sheets, but we use a list to keep the structure
        read_func = lambda _: pd.read_csv(file_path, sep='\t')
    first_digits_count = [0] * 9
    for sheet_name in sheet_names:
        df = read_func(sheet_name)
        values = pd.Series(df.values.ravel())
        digit_series = df.map(extract_first_significant_digit).stack().value_counts().reindex(np.arange(1, 10), fill_value=0)
        first_digits_count += digit_series


        duplicated_values = values[values.duplicated(keep=False)].unique()

        decimal_occurrences = {}
        digit_occurrences = {}
        for val in df.values.ravel():
            first_five = extract_first_five_decimals(val)
            new_digits = extract_digits(val)
            if first_five is not None:
                decimal_occurrences.setdefault(first_five, []).append(val)
            if new_digits is not None:
                digit_occurrences.setdefault(new_digits, []).append(val)

        num_groups = len(set(decimal_occurrences.keys()))
        num_groups_digits = len(set(digit_occurrences.keys()))
        color_gradient_red_yellow = interpolate_colors_red_to_yellow(max(num_groups, num_groups_digits))
        color_gradient_green_blue = interpolate_colors_green_to_blue(max(num_groups, num_groups_digits))

        num_groups_duplicates= len(set(tuple(vals) for vals in decimal_occurrences.values() if len(vals) > 1))
        num_groups_duplicates_digits = len(set(tuple(vals) for vals in digit_occurrences.values() if len(vals) > 1))
        color_gradient_red_yellow_duplicates = interpolate_colors_red_to_yellow(num_groups_duplicates)
        color_gradient_green_blue_duplicates = interpolate_colors_green_to_blue(num_groups_duplicates_digits)

        filled_cells = set()  # Track cells that have been filled

        if len(duplicated_values) > 1 or any(len(vals) > 1 for vals in decimal_occurrences.values()) or any(len(vals) > 1 for vals in digit_occurrences.values()):
            workbook_modified = True
            ws = wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(row)
                if r_idx > 1:
                    for c_idx, cell in enumerate(ws[r_idx]):
                        cell_position = (r_idx, c_idx)  # Track cell position
                        if cell.value in duplicated_values:
                            cell.fill = PatternFill(start_color="add8e6", end_color="add8e6", fill_type="solid")
                        first_five = extract_first_five_decimals(cell.value)
                        if first_five is not None and len(decimal_occurrences[first_five]) > 1:
                            decimal_duplication_found = True
                            color_index = list(decimal_occurrences.keys()).index(first_five)
                            try:
                                fill_color = color_gradient_red_yellow_duplicates[color_index]
                            except:
                                fill_color = color_gradient_red_yellow[color_index % len(color_gradient_red_yellow)]
                            # fill_color = color_gradient_red_yellow[color_index % len(color_gradient_red_yellow)]
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                            filled_cells.add(cell_position)  # Mark this cell as filled
                        new_digits = extract_digits(cell.value)
                        if new_digits is not None and len(digit_occurrences[new_digits]) > 1 and cell_position not in filled_cells:
                            decimal_duplication_found = True
                            color_index = list(digit_occurrences.keys()).index(new_digits)
                            try:
                                fill_color = color_gradient_green_blue_duplicates[color_index]
                            except:
                                fill_color = color_gradient_green_blue[color_index % len(color_gradient_green_blue)]
                            # fill_color = color_gradient_green_blue[color_index % len(color_gradient_green_blue)]
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        if workbook_modified:
            base_name = os.path.basename(file_path)
            new_base_name = os.path.splitext(base_name)[0] + ('_duplicateDecimal' if decimal_duplication_found else '_duplicateCell') + '.xlsx'
            new_file_path = os.path.join(output_dir, new_base_name)
            # new_file_path = os.path.join(output_directory, new_base_name)
            wb.save(new_file_path)
    if not first_digits_count.empty:
        total_data_points = first_digits_count.sum()
        plot_benford_law(first_digits_count.tolist(), total_data_points)


for root, dirs, files in os.walk(directory_path):
    for file in tqdm(files, desc='Processing files'):
        tqdm.write(f"Reading file: {file}")

        if file.endswith(('.xlsx', '.csv', '.txt')):
            file_path = os.path.join(root, file)
            file_size = os.path.getsize(file_path)
            if file_size > 614400:  # Skip files larger than 600 KB
                tqdm.write(f"Skipping file due to size limit: {file}")
                continue

            # Prepare the output directory for this file
            relative_path = os.path.relpath(root, directory_path)
            output_dir = os.path.join(output_directory, relative_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # Process the file
            process_file(file_path, output_dir)

tqdm.write("All files processed.")
